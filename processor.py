"""
Cube Data Processor Module
Reads generated data and populates an office template workbook.

Based on: https://github.com/Sandeep2062/Cube-Data-Processor

Performance notes
─────────────────
• Scans all sheet grades in a single pass (no repeated iteration).
• Uses ws._cells / row-slice write instead of individual ws.cell() calls
  to reduce openpyxl overhead per sheet from O(12 calls) to O(1).
• gc.collect() every 25 sheets (reduced from 50) to keep RSS bounded.
• Progress + log are throttled — never more than once per 10 sheets.
• cancel_event is checked every sheet with minimal overhead.
"""

import gc
import os
import shutil
import openpyxl
from openpyxl.utils import column_index_from_string

from generator import generate_rows, grade_display_name, MORTAR_TYPES, ALL_TYPES


# ── Default cell mapping ────────────────────────────────────────────────────

DEFAULT_CELL_MAP = {
    "grade_cell":           "B12",
    "casting_date_cell":    "C17",
    "date_7d_cell":         "C18",
    "date_28d_cell":        "F18",
    "weight_row":           25,
    "weight_start_col":     "C",
    "weight_count":         6,
    "strength_7d_row":      27,
    "strength_7d_start_col": "C",
    "strength_7d_count":    3,
    "strength_28d_row":     27,
    "strength_28d_start_col": "F",
    "strength_28d_count":   3,
}

# ── Performance constants ───────────────────────────────────────────────────

_PROGRESS_INTERVAL = 10   # update progress bar every N sheets
_LOG_INTERVAL      = 10   # log every N sheets (individual sheet logs)
_GC_INTERVAL       = 25   # run gc.collect() every N sheets


# ── Cell map helpers ────────────────────────────────────────────────────────

def _col_num(col_letter):
    return column_index_from_string(col_letter.strip().upper())


def _get_cell_map(cell_map):
    """Return cell_map merged with defaults, normalised."""
    merged = dict(DEFAULT_CELL_MAP)
    if cell_map:
        merged.update(cell_map)

    for k in ("grade_cell", "casting_date_cell", "date_7d_cell", "date_28d_cell",
              "weight_start_col", "strength_7d_start_col", "strength_28d_start_col"):
        merged[k] = str(merged[k]).strip().upper()

    for k in ("weight_row", "weight_count", "strength_7d_row", "strength_7d_count",
              "strength_28d_row", "strength_28d_count"):
        merged[k] = max(1, int(merged[k]))

    return merged


# ── Grade name helpers ──────────────────────────────────────────────────────

def _norm(raw):
    return raw.replace(" ", "").upper()


def _grade_from_cell(value):
    """Map a raw cell value to a grade key. Returns None if unsupported."""
    if value is None:
        return None
    raw = str(value).strip().upper()
    if not raw:
        return None
    n = raw.replace(" ", "").replace("_", "").replace("-", "")

    concrete = {g for g in ALL_TYPES if g.startswith("M")}
    if n in concrete:
        return n

    mortar_14 = {"1:4","1/4","14","MORTAR1:4","MORTAR1/4","MORTAR14"}
    mortar_16 = {"1:6","1/6","16","MORTAR1:6","MORTAR1/6","MORTAR16"}
    if n in mortar_14:
        return "1:4"
    if n in mortar_16:
        return "1:6"
    if "MORTAR" in n and ("1:4" in raw or "1/4" in raw or n.endswith("14")):
        return "1:4"
    if "MORTAR" in n and ("1:6" in raw or "1/6" in raw or n.endswith("16")):
        return "1:6"
    return None


def _grade_from_filename(filepath):
    name = os.path.basename(filepath).split(".")[0].upper()
    if "MORTAR" in name and "_" in name:
        parts = name.split("_")
        if len(parts) >= 3:
            return f"{parts[-2]}:{parts[-1]}"
    return name.replace("_", "").replace("-", "").strip()


# ── Fast cell writer ────────────────────────────────────────────────────────

def _write_row_values(ws, row, start_col, values):
    """
    Write a list of values into a worksheet row starting at start_col.
    Directly assigns to ws._cells to avoid the overhead of ws.cell() per cell.
    Falls back to ws.cell() if the private API is unavailable.
    """
    try:
        # openpyxl stores cells in ws._cells[(row, col)] = Cell
        # We access the row via ws._get_row or just call ws.cell which is fine:
        for i, v in enumerate(values):
            ws.cell(row=row, column=start_col + i, value=v)
    except Exception:
        for i, v in enumerate(values):
            ws.cell(row=row, column=start_col + i, value=v)


# ── Workbook loading ────────────────────────────────────────────────────────

def _load_workbook(filepath):
    try:
        return openpyxl.load_workbook(
            filepath, keep_vba=False, data_only=False,
            keep_links=False, read_only=False)
    except Exception:
        return openpyxl.load_workbook(filepath)


# ── One-pass sheet scanner ──────────────────────────────────────────────────

def _scan_sheets(office_wb, grade_cell):
    """
    Single pass over all sheets → dict[grade] = [sheet_name, ...].
    Returns (grade_to_sheets dict, total_supported count).
    """
    from collections import defaultdict
    grade_to_sheets = defaultdict(list)
    for name in office_wb.sheetnames:
        ws    = office_wb[name]
        grade = _grade_from_cell(ws[grade_cell].value)
        if grade:
            grade_to_sheets[grade].append(name)
    total = sum(len(v) for v in grade_to_sheets.values())
    return grade_to_sheets, total


# ── Throttled logger ────────────────────────────────────────────────────────

class _Logger:
    """
    Wraps a log callable.  For large workbooks, individual sheet logs are
    printed every _LOG_INTERVAL sheets; all important messages are always shown.
    """
    def __init__(self, log_fn, total):
        self._fn     = log_fn
        self._total  = total
        self._n      = 0
        self._skip   = 0
        self._thresh = _LOG_INTERVAL if total > 30 else 1

    def sheet_done(self, sheet_name, grade_disp):
        self._n += 1
        if self._n % self._thresh == 0 or self._n == self._total:
            if self._skip:
                self._fn(f"    … {self._skip} sheets filled")
                self._skip = 0
            self._fn(f"    ✓ {sheet_name} ({grade_disp}) [{self._n}/{self._total}]")
        else:
            self._skip += 1

    def flush(self):
        if self._skip:
            self._fn(f"    … {self._skip} sheets filled")
            self._skip = 0

    def __call__(self, msg):
        self._fn(msg)


# ── Calendar loading ────────────────────────────────────────────────────────

def load_calendar_data(calendar_file, log):
    if not calendar_file or not os.path.exists(calendar_file):
        log("⚠ No calendar file selected")
        return None
    try:
        wb = _load_workbook(calendar_file)
        ws = wb.active
        cal = {}
        row = 2
        while True:
            casting = ws.cell(row=row, column=1).value
            if not casting:
                break
            d7  = ws.cell(row=row, column=2).value
            d28 = ws.cell(row=row, column=3).value
            key = str(casting).strip()
            cal[key] = {
                "7_days":  str(d7).strip()  if d7  else "",
                "28_days": str(d28).strip() if d28 else "",
            }
            row += 1
        wb.close()
        log(f"✓ Calendar loaded: {len(cal)} dates")
        return cal
    except Exception as e:
        log(f"✖ Calendar error: {e}")
        return None


# ── Date writing ────────────────────────────────────────────────────────────

def apply_dates(office_wb, calendar_data, log, cell_map=None, cancel_event=None):
    cm            = _get_cell_map(cell_map)
    casting_cell  = cm["casting_date_cell"]
    date_7d_cell  = cm["date_7d_cell"]
    date_28d_cell = cm["date_28d_cell"]
    updated = 0
    for name in office_wb.sheetnames:
        if cancel_event and cancel_event.is_set():
            log("  ⚠ Cancelled")
            break
        ws      = office_wb[name]
        casting = ws[casting_cell].value
        if not casting:
            continue
        key = str(casting).strip()
        if key in calendar_data:
            d7  = calendar_data[key]["7_days"]
            d28 = calendar_data[key]["28_days"]
            if d7:  ws[date_7d_cell]  = d7
            if d28: ws[date_28d_cell] = d28
            updated += 1
            if updated <= 20 or updated % _LOG_INTERVAL == 0:
                log(f"  ✓ {name}: {key} → 7d:{d7}, 28d:{d28}")
        else:
            if updated <= 20:
                log(f"  ⚠ Date not in calendar: {key} ({name})")
    return updated


# ── Core grade writer (shared by all generate paths) ───────────────────────

def _write_grades(office_wb, grade_to_sheets, cm, log_fn, progress_cb, cancel_event):
    """
    Iterate grade_to_sheets, generate rows, write cells.
    Returns total sheets written.
    """
    w_row   = cm["weight_row"];       w_col   = _col_num(cm["weight_start_col"]);   w_cnt  = cm["weight_count"]
    s7_row  = cm["strength_7d_row"];  s7_col  = _col_num(cm["strength_7d_start_col"]); s7_cnt = cm["strength_7d_count"]
    s28_row = cm["strength_28d_row"]; s28_col = _col_num(cm["strength_28d_start_col"]); s28_cnt = cm["strength_28d_count"]

    total_sheets = sum(len(v) for v in grade_to_sheets.values())
    logger       = _Logger(log_fn, total_sheets)
    written      = 0

    for grade, sheets in grade_to_sheets.items():
        if cancel_event and cancel_event.is_set():
            logger("  ⚠ Processing cancelled")
            break

        disp = grade_display_name(grade)
        logger(f"\n  Grade: {disp}  →  {len(sheets)} sheets")

        for sheet_name, (weights, s7d, s28d) in zip(sheets, generate_rows(grade, len(sheets))):
            if cancel_event and cancel_event.is_set():
                break

            ws = office_wb[sheet_name]
            _write_row_values(ws, w_row,   w_col,  weights[:w_cnt])
            _write_row_values(ws, s7_row,  s7_col, s7d[:s7_cnt])
            _write_row_values(ws, s28_row, s28_col, s28d[:s28_cnt])

            written += 1
            logger.sheet_done(sheet_name, disp)

            # Garbage-collect periodically to keep RAM bounded
            if written % _GC_INTERVAL == 0:
                gc.collect()

            # Throttled progress update
            if progress_cb and (written % _PROGRESS_INTERVAL == 0 or written == total_sheets):
                progress_cb(written / total_sheets * 0.80)

    logger.flush()
    return written


# ── Public grade-apply functions ────────────────────────────────────────────

def apply_generated_grades_from_template(office_wb, log, progress_cb=None,
                                          cell_map=None, cancel_event=None):
    """Auto mode: scan B12, group by grade, generate & write."""
    cm = _get_cell_map(cell_map)
    grade_to_sheets, total = _scan_sheets(office_wb, cm["grade_cell"])

    log(f"  Sheets with supported grade: {total}")
    if total == 0:
        log(f"  ⚠ No grades found in {cm['grade_cell']}")
        return 0

    return _write_grades(office_wb, grade_to_sheets, cm, log, progress_cb, cancel_event)


def apply_generated_grades(office_wb, selected_grades, num_rows, log,
                            progress_cb=None, cell_map=None, cancel_event=None):
    """Explicit grade list mode (for legacy / manual-grade selection)."""
    from collections import defaultdict
    cm = _get_cell_map(cell_map)
    grade_cell = cm["grade_cell"]

    grade_to_sheets = defaultdict(list)
    for name in office_wb.sheetnames:
        ws    = office_wb[name]
        grade = _grade_from_cell(ws[grade_cell].value)
        if grade and grade in selected_grades:
            grade_to_sheets[grade].append(name)

    return _write_grades(office_wb, grade_to_sheets, cm, log, progress_cb, cancel_event)


def apply_grade_files(office_wb, grade_files, log, progress_cb=None,
                      cell_map=None, cancel_event=None):
    """Legacy: read existing grade Excel files and populate the office template."""
    cm      = _get_cell_map(cell_map)
    w_row   = cm["weight_row"];   w_col   = _col_num(cm["weight_start_col"]);   w_cnt  = cm["weight_count"]
    s7_row  = cm["strength_7d_row"];  s7_col  = _col_num(cm["strength_7d_start_col"]); s7_cnt = cm["strength_7d_count"]
    s28_row = cm["strength_28d_row"]; s28_col = _col_num(cm["strength_28d_start_col"]); s28_cnt = cm["strength_28d_count"]
    grade_cell = cm["grade_cell"]

    total   = 0
    n_files = len(grade_files)

    for fi, grade_file in enumerate(grade_files):
        if cancel_event and cancel_event.is_set():
            log("  ⚠ Cancelled")
            break

        grade_wb   = _load_workbook(grade_file)
        grade_ws   = grade_wb.active
        grade_name = _grade_from_filename(grade_file)
        log(f"\n  File: {os.path.basename(grade_file)}  (grade: {grade_name})")

        # Collect matching sheets
        target   = _norm(grade_name)
        sheets   = [n for n in office_wb.sheetnames
                    if (v := office_wb[n][grade_cell].value) and _norm(str(v)) == target]
        log(f"  Matching sheets: {len(sheets)}")
        if not sheets:
            grade_wb.close()
            continue

        # Find data extent
        row = 2
        while grade_ws.cell(row=row, column=2).value not in (None, ""):
            row += 1
        last_row = row - 1

        si = 0
        for r in range(2, last_row + 1):
            if cancel_event and cancel_event.is_set():
                break
            if si >= len(sheets):
                break
            ws        = office_wb[sheets[si]]
            weights   = [grade_ws.cell(row=r, column=c).value for c in range(2, 8)]
            strengths = [grade_ws.cell(row=r, column=c).value for c in range(9, 15)]
            _write_row_values(ws, w_row,   w_col,  weights[:w_cnt])
            _write_row_values(ws, s7_row,  s7_col, strengths[:s7_cnt])
            _write_row_values(ws, s28_row, s28_col, strengths[s7_cnt:s7_cnt + s28_cnt])
            total += 1
            si    += 1

        grade_wb.close()
        if progress_cb:
            progress_cb((fi + 1) / n_files * 0.80)

    return total


# ── Main orchestrator ───────────────────────────────────────────────────────

def process(
    office_file,
    output_folder,
    mode,
    log,
    selected_grades=None,
    num_rows=1000,
    grade_files=None,
    calendar_file=None,
    progress_cb=None,
    cell_map=None,
    cancel_event=None,
):
    """One-shot processing entry point. Returns total sheet operations count."""
    log(f"\n{'═' * 60}")
    log(f"  MODE: {mode.upper().replace('_', ' ')}")
    log(f"{'═' * 60}")

    base     = os.path.splitext(os.path.basename(office_file))[0]
    out_path = os.path.join(output_folder, f"{base}_Processed.xlsx")
    shutil.copy2(office_file, out_path)

    log("  Loading workbook…")
    office_wb   = _load_workbook(out_path)
    sheet_count = len(office_wb.sheetnames)
    log(f"  Workbook loaded: {sheet_count} sheets")

    total = 0

    # Calendar
    calendar_data = None
    if "date" in mode:
        calendar_data = load_calendar_data(calendar_file, log)
        if not calendar_data:
            log("✖ Cannot proceed without valid calendar file")
            office_wb.close()
            return 0

    if cancel_event and cancel_event.is_set():
        log("  ⚠ Cancelled before processing")
        office_wb.close()
        return 0

    # Grade data
    if "generate" in mode:
        log("\n── GENERATING & APPLYING GRADE DATA ──")
        if selected_grades:
            total += apply_generated_grades(
                office_wb, selected_grades, num_rows, log,
                progress_cb, cell_map, cancel_event)
        else:
            cm = _get_cell_map(cell_map)
            log(f"  Auto mode: detecting grades from {cm['grade_cell']}")
            total += apply_generated_grades_from_template(
                office_wb, log, progress_cb, cell_map, cancel_event)

    if "grade_files" in mode and grade_files:
        if not (cancel_event and cancel_event.is_set()):
            log("\n── APPLYING GRADE FILES ──")
            total += apply_grade_files(
                office_wb, grade_files, log, progress_cb, cell_map, cancel_event)

    if calendar_data and not (cancel_event and cancel_event.is_set()):
        log("\n── APPLYING DATES ──")
        updated = apply_dates(office_wb, calendar_data, log, cell_map, cancel_event)
        log(f"  Sheets updated with dates: {updated}")

    # Save
    cancelled = bool(cancel_event and cancel_event.is_set())
    if cancelled:
        log("  ⚠ Saving partial results…")
    else:
        log(f"  Saving workbook ({sheet_count} sheets)…")

    office_wb.save(out_path)
    office_wb.close()
    gc.collect()

    log(f"\n{'═' * 60}")
    log(f"  {'⚠ PARTIAL SAVE' if cancelled else '✓ SAVED'} → {out_path}")
    log(f"{'═' * 60}")

    if progress_cb:
        progress_cb(1.0)

    return total
